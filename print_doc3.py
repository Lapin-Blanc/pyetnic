import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pyetnic.services import lire_document_3
from pyetnic.services.models import OrganisationId
from pyetnic.config import Config
from pyetnic.log_config import configure_logging
from datetime import datetime, timedelta

# Configuration du logging
configure_logging()
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

def generate_excel(orga_id):
    doc3 = lire_document_3(orga_id)
    
    if not doc3 or not doc3.activiteListe:
        logger.warning("Aucune activité d'enseignement trouvée dans le document 3")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attributions"
    
    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    centered = Alignment(horizontal='center')
    
    # En-tête
    headers = ["N°", "Branche", "Catégorie", "Nom Branche", "Année", "Périodes Doc8", "Périodes Prévues Doc2", 
               "Périodes Réelles Doc2", "Enseignant", "Matricule", "Statut", "Disponibilité", "Périodes Attribuées"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
    
    # Données
    row_idx = 2
    for act_idx, activite in enumerate(doc3.activiteListe.activite):
        # Si l'activité a des enseignants
        if activite.enseignantListe and activite.enseignantListe.enseignant:
            for ens_idx, enseignant in enumerate(activite.enseignantListe.enseignant):
                ws.cell(row=row_idx, column=1, value=activite.coNumBranche)
                ws.cell(row=row_idx, column=2, value=activite.coCategorie)
                ws.cell(row=row_idx, column=3, value=activite.coCategorie)
                ws.cell(row=row_idx, column=4, value=activite.teNomBranche)
                ws.cell(row=row_idx, column=5, value=activite.noAnneeEtude)
                ws.cell(row=row_idx, column=6, value=activite.nbPeriodesDoc8)
                ws.cell(row=row_idx, column=7, value=activite.nbPeriodesPrevuesDoc2)
                ws.cell(row=row_idx, column=8, value=activite.nbPeriodesReellesDoc2)
                ws.cell(row=row_idx, column=9, value=f"{enseignant.teNomEns} {enseignant.tePrenomEns}")
                ws.cell(row=row_idx, column=10, value=enseignant.noMatEns)
                ws.cell(row=row_idx, column=11, value=enseignant.teStatut)
                ws.cell(row=row_idx, column=12, value=enseignant.coDispo if enseignant.coDispo else "")
                ws.cell(row=row_idx, column=13, value=enseignant.nbPeriodesAttribuees)
                row_idx += 1
        else:
            # Si l'activité n'a pas d'enseignants
            ws.cell(row=row_idx, column=1, value=activite.coNumBranche)
            ws.cell(row=row_idx, column=2, value=activite.coCategorie)
            ws.cell(row=row_idx, column=3, value=activite.coCategorie)
            ws.cell(row=row_idx, column=4, value=activite.teNomBranche)
            ws.cell(row=row_idx, column=5, value=activite.noAnneeEtude)
            ws.cell(row=row_idx, column=6, value=activite.nbPeriodesDoc8)
            ws.cell(row=row_idx, column=7, value=activite.nbPeriodesPrevuesDoc2)
            ws.cell(row=row_idx, column=8, value=activite.nbPeriodesReellesDoc2)
            row_idx += 1
    
    # Ajustement de la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarde du fichier
    filename = f"doc3_attributions_{orga_id.anneeScolaire}_{orga_id.etabId}_{orga_id.numAdmFormation}_{orga_id.numOrganisation}.xlsx"
    wb.save(filename)
    logger.info(f"Fichier Excel généré : {filename}")

def get_current_school_year():
    now = datetime.now()
    current_year = now.year
    
    # Trouver le premier vendredi de juillet
    july_first = datetime(current_year, 7, 1)
    days_until_friday = (4 - july_first.weekday() + 7) % 7
    first_friday_july = july_first + timedelta(days=days_until_friday)
    
    if now < first_friday_july:
        return f"{current_year-1}-{current_year}"
    else:
        return f"{current_year}-{current_year+1}"

def get_user_input():
    default_year = get_current_school_year()
    annee_scolaire = input(f"Entrez l'année scolaire (par défaut: {default_year}): ") or default_year
    etab_id = int(input(f"Entrez l'identifiant FASE de l'établissement (par défaut: {Config.ETAB_ID}): ") or Config.ETAB_ID)
    num_admin = int(input("Entrez le numéro administratif de la formation: "))
    num_organisation = int(input("Entrez le numéro d'organisation: "))
    return annee_scolaire, etab_id, num_admin, num_organisation

if __name__ == "__main__":
    annee_scolaire, etab_id, num_admin, num_organisation = get_user_input()
    
    orga_id = OrganisationId(
        anneeScolaire=annee_scolaire,
        etabId=etab_id,
        numAdmFormation=num_admin,
        numOrganisation=num_organisation
    )
    
    generate_excel(orga_id)
