import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pyetnic.services import lire_document_2
from pyetnic.services.models import OrganisationId
from pyetnic.config import Config
from pyetnic.log_config import configure_logging
from datetime import datetime, timedelta

# Configuration du logging
configure_logging()
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

def generate_excel(orga_id):
    doc2 = lire_document_2(orga_id)
    
    if not doc2 or not doc2.activiteEnseignementDetail or not doc2.activiteEnseignementDetail.activiteEnseignementListe:
        logger.warning("Aucune activité d'enseignement trouvée dans le document 2")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Activités d'enseignement"
    
    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    centered = Alignment(horizontal='center')
    
    # En-tête
    headers = ["N°", "Branche", "Catégorie", "Nom", "Année", "Élèves", "Périodes Branche", "Prévues An1", "Prévues An2", "Réelles An1", "Réelles An2"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
    
    # Données
    for idx, activite in enumerate(doc2.activiteEnseignementDetail.activiteEnseignementListe.activiteEnseignement, 2):
        ws.cell(row=idx, column=1, value=idx-1)
        ws.cell(row=idx, column=2, value=activite.coNumBranche)
        ws.cell(row=idx, column=3, value=activite.coCategorie)
        ws.cell(row=idx, column=4, value=activite.teNomBranche)
        ws.cell(row=idx, column=5, value=activite.coAnnEtude)
        ws.cell(row=idx, column=6, value=activite.nbEleveC1)
        ws.cell(row=idx, column=7, value=activite.nbPeriodeBranche)
        ws.cell(row=idx, column=8, value=activite.nbPeriodePrevueAn1)
        ws.cell(row=idx, column=9, value=activite.nbPeriodePrevueAn2)
        ws.cell(row=idx, column=10, value=activite.nbPeriodeReelleAn1)
        ws.cell(row=idx, column=11, value=activite.nbPeriodeReelleAn2)
    
    # Ajustement de la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarde du fichier
    filename = f"doc2_organisation_{orga_id.anneeScolaire}_{orga_id.etabId}_{orga_id.numAdmFormation}_{orga_id.numOrganisation}.xlsx"
    wb.save(filename)
    logger.info(f"Fichier Excel généré : {filename}")

def get_current_school_year():
    now = datetime.now()
    current_year = now.year
    
    # Trouver le premier vendredi de juillet
    july_first = datetime(current_year, 7, 1)
    days_until_friday = (4 - july_first.weekday() + 7) % 7
    first_friday_july = july_first + timedelta(days=days_until_friday)
    
    if now < first_friday_july:
        return f"{current_year-1}-{current_year}"
    else:
        return f"{current_year}-{current_year+1}"

def get_user_input():
    default_year = get_current_school_year()
    annee_scolaire = input(f"Entrez l'année scolaire (par défaut: {default_year}): ") or default_year
    num_admin = int(input("Entrez le numéro administratif de la formation: "))
    num_organisation = int(input("Entrez le numéro d'organisation: "))
    return annee_scolaire, num_admin, num_organisation

if __name__ == "__main__":
    annee_scolaire, num_admin, num_organisation = get_user_input()
    
    orga_id = OrganisationId(
        anneeScolaire=annee_scolaire,
        etabId=Config.ETAB_ID,
        numAdmFormation=num_admin,
        numOrganisation=num_organisation
    )
    
    generate_excel(orga_id)